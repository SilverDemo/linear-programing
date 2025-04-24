#from cvxopt import matrix, solvers
import numpy as np
from openpyxl import load_workbook
import argparse
from math import inf
from pulp import *
import sys

def read_exel(excel_path):
    workbook = load_workbook(filename=excel_path, data_only=True)
    ret = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data = [[cell if cell is not None else "None"  for cell in row] for row in sheet.iter_rows(values_only=True)]
        np_array = np.array(data)
        ret[sheet_name] = np_array
    return ret
    
def calculate(price_break, demands, demand_nodes, modes, supplier_nodes, capacity, holding_cost_per_unit, costs, fixed_cost, time_lead_cost):
    problem = LpProblem("Transportation_Problem", LpMinimize)

    # Decision variables
    purschase_quantity_per_supplier = LpVariable.dicts("purschase_quantity_per_supplier", indices=((sup) for sup in supplier_nodes) , lowBound=0, cat=const.LpInteger)
    transport_mode_binary = LpVariable.dicts("transport_mode_binary", indices=((sup, dest, mode) for sup in supplier_nodes for dest in demand_nodes for mode in modes), cat=const.LpBinary) #(9)
    purschase_quantity_per_mode = LpVariable.dicts("purschase_quantity_per_mode", indices=((sup, dest, mode) for sup in supplier_nodes for dest in demand_nodes for mode in modes),lowBound=0, cat=const.LpInteger) #(8)
    price_break_binary = LpVariable.dicts("price_break_binary", indices=((sup, level) for sup in supplier_nodes for level in price_break[sup].keys()), cat=const.LpBinary) #(5)
    # purschase_quantity_per_supplier_linearized = price_break_binary * purschase_quantity_per_supplier
    purschase_quantity_per_supplier_linearized = LpVariable.dicts("purschase_quantity_per_supplier_linearized", indices=((sup, level) for sup in supplier_nodes for level in price_break[sup].keys()), lowBound=0, cat=const.LpInteger)


    # Objective function
    purchase_cost = lpSum([purschase_quantity_per_supplier_linearized[i, r] * int(price_break[i][r][2])] for i in price_break.keys() for r in price_break[i].keys())
    transport_cost = lpSum([purschase_quantity_per_mode[i,dest,mode] * costs[i,dest,mode] + fixed_cost[i,dest,mode] * transport_mode_binary[i,dest,mode]] for i in supplier_nodes for dest in demand_nodes for mode in modes)
    holding_cost = lpSum([holding_cost_per_unit * time_lead_cost[i,dest,mode] * purschase_quantity_per_mode[i,dest,mode]]  for i in supplier_nodes for dest in demand_nodes for mode in modes)
    problem += purchase_cost + transport_cost + holding_cost # (1)

    #constraint
    # linearize purschase_quantity_per_supplier_linearized = purschase_quantity_per_supplier * price_break_binary
    for i in supplier_nodes:
        problem += lpSum([purschase_quantity_per_supplier_linearized[i,r] for r in price_break[i].keys()]) == purschase_quantity_per_supplier[i]
        for r in price_break[i].keys():
            # eliminate products (M = (int)(capacity[i]))
            problem += purschase_quantity_per_supplier_linearized[i, r] >= purschase_quantity_per_supplier[i] - (int)(capacity[i]) * (1 - price_break_binary[i, r])
            # problem += purschase_quantity_per_supplier_linearized[i, r] >= 0 # lowbound

            # price break bounds check
            problem += purschase_quantity_per_supplier_linearized[i,r] <= int(price_break[i][r][1]) # max check (4)
            problem += purschase_quantity_per_supplier_linearized[i,r] >= int(price_break[i][r][0]) * price_break_binary[i, r] # min check (4)

            # price_break_binary = 1 if purschase_quantity_per_supplier_linearized > 0
            problem += price_break_binary[i,r] * max(int(capacity[i]), 1) >= purschase_quantity_per_supplier_linearized[i, r] #(3)
            problem += price_break_binary[i, r] <= purschase_quantity_per_supplier_linearized[i, r] #(3)

 

    # purschase_quantity_per_supplier - purschase_quantity_per_mode constraint
    total_quantity_per_mode = {}
    for i in supplier_nodes:
        total_quantity_per_mode[i] = lpSum([purschase_quantity_per_mode[i,dest,mode]] for dest in demand_nodes for mode in modes)
        problem += (total_quantity_per_mode[i] == purschase_quantity_per_supplier[i]) #(2)

    # transport_mode_binary constraint
    for i in supplier_nodes:
        for dest in demand_nodes:
            for mode in modes:
                # problem += purschase_quantity_per_mode[i,dest,mode] >= transport_mode_binary[i,dest,mode]
                problem += transport_mode_binary[i, dest, mode] * max(demands[dest], 1) >= purschase_quantity_per_mode[i, dest, mode]
                problem += transport_mode_binary[i, dest, mode] <= purschase_quantity_per_mode[i, dest, mode]
                """
                transport_mode_binary | purschase_quantity_per_mode | 1 | 2
                ------------------------------------------------------------
                0  | > 0 | false | true
                0  | 0   | true | true
                1  | > 0 | true | true
                1  | 0   | true | false
                """

    # capacity constraint
    for i in supplier_nodes:
        problem += purschase_quantity_per_supplier[i] <= int(capacity[i]) #(6)

    # demand constraint
    for dest in demand_nodes:
        problem += lpSum(purschase_quantity_per_mode[i,dest,mode] for mode in modes for i in supplier_nodes) == demands[dest] #(7)

    problem.solve()

    print("Status:", LpStatus[problem.status])
    print("Optimal Solution:")
    print(problem.objective.value())
    for v in problem.variables():
        print("{} = {}". format(v.name, v.varValue))

    problem.toJson("probSum.json")
    print("Optimal Value (Objective Function):", value(problem.objective))


def solutionDict(data):
    result_dict = {}

    for item in data:
        parts = item.split(" = ")
        key_parts = parts[0].split("_('_")[1].split("', '")
        key = tuple(key_parts[:-1])  # Use a tuple for multi-part keys
        value = float(parts[1])
        if key not in result_dict:
            result_dict[key] = {}
        result_dict[key][key_parts[-1]] = value
    return result_dict


def get_cost(array, supply, demand, mode):
    row = None
    for s in array:
        if s[0] == supply:
            row = s
    if row is None:
        return 0
    modes = array[0]
    demands = array[1]
    no_demand = demands.shape
    for i in range(no_demand[0]):
        if modes[i] == mode and demands[i] == demand:
            return float(row[i])
    return 0
   
def debugData(data, name=None, tab = 0):
    if name is not None:
        print("Data {}: ".format(name))

    if isinstance(data, dict):
        debugDict(data, tab=tab + 1)
    elif isinstance(data, list):
        debugList(data, tab=tab+1)
    else:
        print("{}{}".format(tab* "\t", data))

def debugDict(data, tab = 0):
    for key, value in data.items():
        if isinstance(value, dict):
            print("{}{}: ".format(tab*"\t", key))
            debugDict(value, tab=tab+1)
        elif isinstance(value, list):
            print("{}{}: ".format(tab*"\t", key))
            debugList(data, tab=tab+1)
        else:
            print("{}{}: {}".format(tab*"\t", key, value))

def debugList(data, tab = 0):
    print("{}[".format(tab*"\t"))
    for value in data:
        if isinstance(value, dict):
            print("{}".format(tab*"\t"), end="")
            debugDict(value, tab=tab+1)
        elif isinstance(value, list):
            print("{}".format(tab*"\t"), end="")
            debugList(data, tab=tab+1)
        else:
            print("{}{}".format((tab+1)*"\t", value))
    print("{}]".format(tab*"\t"))

    
def main():
    parser = argparse.ArgumentParser(description="Read Excel file and print its path")
    parser.add_argument('excel_file', type=str, help="Path to the Excel file")
    args = parser.parse_args()
    
    # parse excel
    input_values = read_exel(args.excel_file)
    
    demand_nodes = list(input_values["demand"][:, 0])
    demands = {node: int(demand) for node, demand in input_values["demand"]}

    max_quantity = sum([de for _, de in demands.items()])

    price_break = {}
    # price_break = {(sup, level, min_quan): price for sup, level, min_quan, price in price_break}
    for sup, level, min_quan, max_quan, price in input_values["price_break"][1:]:
        if sup not in price_break.keys():
            price_break[sup] = {}
        if max_quan == "inf":
            max_quan = max_quantity
        price_break[sup][level] = (min_quan, max_quan, price) 


    modes = list(set(input_values["cost"][0, 1:]))

    supplier_nodes = input_values["capacity"][1:]
    supplier_nodes = [ node for node, _ in supplier_nodes]

    capacity = input_values["capacity"][1:]
    capacity = { node : cap for node, cap in capacity}

    holding_cost_per_unit =  input_values["holding_cost_per_unit"][0][0]

    costs = {}
    fixed_cost = {}
    time_lead_cost = {}

    for sup in supplier_nodes:
        for de in demand_nodes:
            for mode in modes:
                costs[(sup, de, mode)] = get_cost(input_values["cost"], sup, de, mode)
                fixed_cost[(sup, de, mode)] = get_cost(input_values["fixed_cost"], sup, de, mode)
                time_lead_cost[(sup, de, mode)] = get_cost(input_values["time_lead_cost"], sup, de, mode)

    # debugData(price_break, name="price_break")
    # debugData(demands, name="demands")
    # debugData(demand_nodes, name="demand_nodes")
    # debugData(modes, name="modes")
    # debugData(supplier_nodes, name="supplier_nodes")
    # debugData(capacity, name="capacity")
    # debugData(holding_cost_per_unit, name="holding_cost_per_unit")
    # debugData(costs, name="costs")
    # debugData(fixed_cost, name="fixed_cost")
    # debugData(time_lead_cost, name="time_lead_cost")
    calculate(price_break, demands, demand_nodes, modes, supplier_nodes, capacity, holding_cost_per_unit, costs, fixed_cost, time_lead_cost)



    

    
if __name__ == "__main__":
    main()
