import os
from openpyxl import load_workbook
import openpyxl
import argparse
from pulp import *
import sys
import numpy as np
from math import inf
from pathlib import Path
import json
import inspect
import copy
import locale
locale.setlocale(locale.LC_ALL, 'en_US')

class _Value:
	def __init__(self, a):
		self.a = a
	def value(self):
		return self.a
def retrieve_name(var):
	"""
	Gets the name of var. Does it from the out most frame inner-wards.
	:param var: variable to get name from.
	:return: string
	"""
	for fi in reversed(inspect.stack()):
		names = [var_name for var_name, var_val in fi.frame.f_locals.items() if var_val is var]
		if len(names) > 0:
			return names[0]

class Watcher:
	watching = dict()
	def __call__(self, *args, **kwargs):
		for i in args:
			self.watching[retrieve_name(i)] = i
		for k, v in kwargs.items():
			self.watching[k] = v
	def __getitem__(self, key):
		return self.watching[key]

	def keys(self):
		return self.watching.keys()

def format2dTable(dict_data):
	col_label = list(dict_data.keys())
	row_label = list(dict_data[col_label[0]].keys())

	form = "|{:<16}" * (len(col_label) + 1)
	ret = form.format(" ", *col_label)
	for row in row_label:
		row_Data = ( v[row] for col, v in dict_data.items())
		ret += "\n"
		ret += form.format(row, *row_Data)
	return ret

def format1dTable(dict_data):
	ret = ""
	form = "{:<16} : {} \n"
	for k,v in dict_data.items():
		ret += form.format(k, v)
	return ret

def format3dTable(dict_data):
	ret = ""
	form = "{}:\n{}\n\n"
	for k,v in dict_data.items():
		ret += form.format(k, format2dTable(v))
	return ret


class ProblemParam:
	def __init__(self):
		self.product_groups = set()
		self.factories = set()
		self.time_periods = set()
		self.r_i = dict()
		self.b_i = dict()
		self.TR_t = dict()
		self.IR_j0 = dict()
		self.Max_P_ij = dict()
		self.IN_ij0 = dict()
		self.BO_ij0 = dict()
		# self.x_ij = dict()
		self.HD_C_ij = dict()
		self.P_C_ij = dict()
		self.T_C_jk = dict()
		self.D_tij = dict()
		self.Fix_ij = dict()
		self.v_ij = dict()
		self.s_ij = dict()

	def toExcel(self, filepath):
		wb = openpyxl.Workbook()
		wb.create_sheet("r_i")
		data = { (i, "value") : _Value(self.r_i[i]) for i in self.product_groups}
		populate_table(wb["r_i"], (1,1), data, self.product_groups, ["value"])
		wb.create_sheet("b_i")
		data = { (i, "value") : _Value(self.b_i[i]) for i in self.product_groups}
		populate_table(wb["b_i"], (1,1), data, self.product_groups, ["value"])

		wb.create_sheet("TR_t")
		data = { (t, "value") : _Value(self.TR_t[t]) for t in self.time_periods}
		populate_table(wb["TR_t"], (1,1), data, self.time_periods, ["value"])
		wb.create_sheet("IR_j0")
		data = { (j, "value") : _Value(self.IR_j0[j]) for j in self.factories}
		populate_table(wb["IR_j0"], (1,1), data, self.factories, ["value"])

		wb.create_sheet("Max_P_ij")
		data = { (i, j) : _Value(self.Max_P_ij[i][j]) for i in self.product_groups for j in self.factories}
		populate_table(wb["Max_P_ij"], (1,1), data, self.product_groups, self.factories)
		wb.create_sheet("IN_ij0")
		data = { (i, j) : _Value(self.IN_ij0[i][j]) for i in self.product_groups for j in self.factories}
		populate_table(wb["IN_ij0"], (1,1), data, self.product_groups, self.factories)
		wb.create_sheet("BO_ij0")
		data = { (i, j) : _Value(self.BO_ij0[i][j]) for i in self.product_groups for j in self.factories}
		populate_table(wb["BO_ij0"], (1,1), data, self.product_groups, self.factories)
		wb.create_sheet("HD_C_ij")
		data = { (i, j) : _Value(self.HD_C_ij[i][j]) for i in self.product_groups for j in self.factories}
		populate_table(wb["HD_C_ij"], (1,1), data, self.product_groups, self.factories)
		wb.create_sheet("P_C_ij")
		data = { (i, j) : _Value(self.P_C_ij[i][j]) for i in self.product_groups for j in self.factories}
		populate_table(wb["P_C_ij"], (1,1), data, self.product_groups, self.factories)
		wb.create_sheet("T_C_jk")
		data = { (j, k) : _Value(self.T_C_jk[j][k]) for j in self.factories for k in self.factories}
		populate_table(wb["T_C_jk"], (1,1), data, self.factories, self.factories)
		wb.create_sheet("D_tij")
		data = { (t, i, j) : _Value(self.D_tij[t][i][j]) for i in self.product_groups for j in self.factories for t in self.time_periods}
		populate_table3d(wb["D_tij"], (1,1), data, self.time_periods, self.product_groups, self.factories)
		wb.create_sheet("Fix_ij")
		data = { (i, j) : _Value(self.Fix_ij[i][j]) for i in self.product_groups for j in self.factories}
		populate_table(wb["Fix_ij"], (1,1), data, self.product_groups, self.factories)
		wb.create_sheet("v_ij")
		data = { (i, j) : _Value(self.v_ij[i][j]) for i in self.product_groups for j in self.factories}
		populate_table(wb["v_ij"], (1,1), data, self.product_groups, self.factories)
		wb.create_sheet("s_ij")
		data = { (i, j) : _Value(self.s_ij[i][j]) for i in self.product_groups for j in self.factories for t in self.time_periods}
		populate_table(wb["s_ij"], (1,1), data, self.product_groups, self.factories)
		wb.save(filepath)

	def parse(self, excel_data):
		self.product_groups = set(excel_data["sets"][1:, 0])
		self.product_groups.discard(None)
		self.product_groups.discard("None")
		self.product_groups = set([i.lower() for i in self.product_groups])

		self.factories = set(excel_data["sets"][1:, 1])
		self.factories.discard(None)
		self.factories.discard("None")
		self.factories = set([j.lower() for j in self.factories])

		self.time_periods = set(excel_data["sets"][1:, 2])
		self.time_periods.discard(None)
		self.time_periods.discard("None")
		self.time_periods = {int(i) for i in self.time_periods}

		self.r_i = { row[0].lower(): float(row[1]) for row in excel_data["r_i"]}
		ProblemParam._check("r_i", set(self.r_i.keys()), self.product_groups)

		self.b_i = { row[0].lower(): float(row[1]) for row in excel_data["b_i"]}
		ProblemParam._check("b_i", set(self.b_i.keys()), self.product_groups)

		self.TR_t = { int(row[0]): float(row[1]) for row in excel_data["TR_t"]}
		ProblemParam._check("TR_t", set(self.TR_t.keys()), self.time_periods)

		self.IR_j0 = { row[0].lower(): float(row[1]) for row in excel_data["IR_j0"].transpose()}
		ProblemParam._check("IR_j0", set(self.IR_j0.keys()), self.factories)

		col_lable, row_lable, self.Max_P_ij = ProblemParam.parse_table(excel_data["Max_P_ij"])
		ProblemParam._check("Max_P_ij column", set(col_lable), self.product_groups)
		ProblemParam._check("Max_P_ij row", set(row_lable), self.factories)

		col_lable, row_lable, self.IN_ij0 = ProblemParam.parse_table(excel_data["IN_ij0"])
		ProblemParam._check("IN_ij0 column", set(col_lable), self.product_groups)
		ProblemParam._check("IN_ij0 row", set(row_lable), self.factories)

		col_lable, row_lable, self.BO_ij0 = ProblemParam.parse_table(excel_data["BO_ij0"])
		ProblemParam._check("BO_ij0 column", set(col_lable), self.product_groups)
		ProblemParam._check("BO_ij0 row", set(row_lable), self.factories)

		# col_lable, row_lable, self.x_ij = ProblemParam.parse_table(excel_data["x_ij"])
		# ProblemParam._check("x_ij column", set(col_lable), self.product_groups)
		# ProblemParam._check("x_ij row", set(row_lable), self.factories)

		col_lable, row_lable, self.HD_C_ij = ProblemParam.parse_table(excel_data["HD_C_ij"])
		ProblemParam._check("HD_C_ij column", set(col_lable), self.product_groups)
		ProblemParam._check("HD_C_ij row", set(row_lable), self.factories)

		col_lable, row_lable, self.P_C_ij = ProblemParam.parse_table(excel_data["P_C_ij"])
		ProblemParam._check("P_C_ij column", set(col_lable), self.product_groups)
		ProblemParam._check("P_C_ij row", set(row_lable), self.factories)

		col_lable, row_lable, self.T_C_jk = ProblemParam.parse_table(excel_data["T_C_jk"])
		ProblemParam._check("T_C_jk column", set(col_lable), self.factories)
		ProblemParam._check("T_C_jk row", set(row_lable), self.factories)

		col_lable, row_lable, self.Fix_ij = ProblemParam.parse_table(excel_data["fix"])
		ProblemParam._check("Fix_ij  column", set(col_lable), self.product_groups)
		ProblemParam._check("Fix_ij  row", set(row_lable), self.factories)

		col_lable, row_lable, self.v_ij = ProblemParam.parse_table(excel_data["v_ij"])
		ProblemParam._check("v_ij  column", set(col_lable), self.product_groups)
		ProblemParam._check("v_ij  row", set(row_lable), self.factories)
		
		col_lable, row_lable, self.s_ij = ProblemParam.parse_table(excel_data["s_ij"])
		ProblemParam._check("s_ij  column", set(col_lable), self.product_groups)
		ProblemParam._check("s_ij  row", set(row_lable), self.factories)



		for t in self.time_periods:
			col_lable, row_lable, self.D_tij[t] = ProblemParam.parse_table(excel_data["D_ij"+str(t)])
			ProblemParam._check("D_ij{} column".format(t), set(col_lable), self.product_groups)
			ProblemParam._check("D_ij{} row".format(t), set(row_lable), self.factories)

	@staticmethod
	def table2dCorrectionFunction(table2d, correct_func):
		for r in table2d.keys():
			for c in table2d[r].keys():
				table2d[r][c] = correct_func(table2d[r][c])

	@staticmethod
	def table3dCorrectionFunction(table3d, correct_func):
		for _1 in table3d.keys():
			for _2 in table3d[_1].keys():
				for _3 in table3d[_1][_2].keys():
					table3d[_1][_2][_3] = correct_func(table3d[_1][_2][_3])

	def parameterCorection(self):
		def correctionFunction(data):
			if data is None or data.strip() == "None" or data.strip() == "":
				return 0
			if data.strip() == "M":
				return inf
			return float(data)

		class collectorForM:
			def __init__(self):
				self.value = 0

			def __call__(self, value):
				if value == inf:
					return value

				self.value += abs(value)
				return value

		class MCorrector:
			def __init__(self, collector):
				self.value = collector.value

			def __call__(self, data):
				if data == inf:
					return self.value
				return data

		ProblemParam.table2dCorrectionFunction(self.Max_P_ij, correctionFunction)
		collector = collectorForM()
		ProblemParam.table2dCorrectionFunction(self.Max_P_ij, collector)
		corrector = MCorrector(collector)
		ProblemParam.table2dCorrectionFunction(self.Max_P_ij, corrector)
		ProblemParam.table2dCorrectionFunction(self.IN_ij0, correctionFunction)
		collector = collectorForM()
		ProblemParam.table2dCorrectionFunction(self.IN_ij0, collector)
		corrector = MCorrector(collector)
		ProblemParam.table2dCorrectionFunction(self.IN_ij0, corrector)
		ProblemParam.table2dCorrectionFunction(self.BO_ij0, correctionFunction)
		collector = collectorForM()
		ProblemParam.table2dCorrectionFunction(self.BO_ij0, collector)
		corrector = MCorrector(collector)
		ProblemParam.table2dCorrectionFunction(self.BO_ij0, corrector)
		# ProblemParam.table2dCorrectionFunction(self.x_ij, correctionFunction)
		# collector = collectorForM()
		# ProblemParam.table2dCorrectionFunction(self.x_ij, collector)
		# corrector = MCorrector(collector)
		# ProblemParam.table2dCorrectionFunction(self.x_ij, corrector)
		ProblemParam.table2dCorrectionFunction(self.HD_C_ij, correctionFunction)
		collector = collectorForM()
		ProblemParam.table2dCorrectionFunction(self.HD_C_ij, collector)
		corrector = MCorrector(collector)
		ProblemParam.table2dCorrectionFunction(self.HD_C_ij, corrector)
		ProblemParam.table2dCorrectionFunction(self.P_C_ij, correctionFunction)
		collector = collectorForM()
		ProblemParam.table2dCorrectionFunction(self.P_C_ij, collector)
		corrector = MCorrector(collector)
		ProblemParam.table2dCorrectionFunction(self.P_C_ij, corrector)
		ProblemParam.table2dCorrectionFunction(self.T_C_jk, correctionFunction)
		collector = collectorForM()
		ProblemParam.table2dCorrectionFunction(self.T_C_jk, collector)
		corrector = MCorrector(collector)
		ProblemParam.table2dCorrectionFunction(self.T_C_jk, corrector)
		ProblemParam.table3dCorrectionFunction(self.D_tij, correctionFunction)
		collector = collectorForM()
		ProblemParam.table3dCorrectionFunction(self.D_tij, collector)
		corrector = MCorrector(collector)
		ProblemParam.table3dCorrectionFunction(self.D_tij, corrector)
		ProblemParam.table2dCorrectionFunction(self.Fix_ij, correctionFunction)
		collector = collectorForM()
		ProblemParam.table2dCorrectionFunction(self.Fix_ij, collector)
		corrector = MCorrector(collector)
		ProblemParam.table2dCorrectionFunction(self.Fix_ij, corrector)
		ProblemParam.table2dCorrectionFunction(self.v_ij, correctionFunction)
		collector = collectorForM()
		ProblemParam.table2dCorrectionFunction(self.v_ij, collector)
		corrector = MCorrector(collector)
		ProblemParam.table2dCorrectionFunction(self.v_ij, corrector)
		ProblemParam.table2dCorrectionFunction(self.s_ij, correctionFunction)
		collector = collectorForM()
		ProblemParam.table2dCorrectionFunction(self.s_ij, collector)
		corrector = MCorrector(collector)
		ProblemParam.table2dCorrectionFunction(self.s_ij, corrector)
	@staticmethod
	def parse_table(excel_table):
		ret = {}
		col_label = [ s.lower() for s in excel_table[1:,0]]
		row_label = [ s.lower() for s in excel_table[0, 1:]]
		for col in range(0, len(col_label)):
			for row in range(0, len(row_label)):
				if col_label[col] not in ret.keys():
					ret[col_label[col]]=dict()
				ret[col_label[col]][row_label[row]] = excel_table[col + 1][row + 1]
		return col_label, row_label, ret



	@staticmethod
	def _check(name, target, crit):
		if target == crit:
			print("{} good".format(name))
		else:
			print("{} bad keys: {}".format(name, target))
			print("{} expect: {}".format(name, crit))
			exit(-1)

	def __str__(self):
		return """
		Product group: {}
		Factories : {}
		Time periods: {}
		r_i : ------------------------------------------------------------------
{}
		b_i : ------------------------------------------------------------------
{}
		TR_t : ------------------------------------------------------------------
{}
		IR_j0 : ------------------------------------------------------------------
{}
		Max_P_ij : ------------------------------------------------------------------
{}
		IN_ij0 : ------------------------------------------------------------------
{}
		BO_ij0 : ------------------------------------------------------------------
{}
		HD_C_ij : ------------------------------------------------------------------
{}
		P_C_ij : ------------------------------------------------------------------
{}
		T_C_jk : ------------------------------------------------------------------
{}
		D_tij : ------------------------------------------------------------------
{}
		Fix_ij : ------------------------------------------------------------------
{}
		v_ij : ------------------------------------------------------------------
{}
		s_ij : ------------------------------------------------------------------
{}
		{}
		""".format(
			self.product_groups, 
			self.factories, 
			self.time_periods, 
			format1dTable(self.r_i),
			format1dTable(self.b_i),
			format1dTable(self.TR_t),
			format1dTable(self.IR_j0),
			format2dTable(self.Max_P_ij),
			format2dTable(self.IN_ij0),
			format2dTable(self.BO_ij0),
			format2dTable(self.HD_C_ij),
			format2dTable(self.P_C_ij),
			format2dTable(self.T_C_jk),
			format3dTable(self.D_tij),
			format2dTable(self.Fix_ij),
			format2dTable(self.v_ij),
			format2dTable(self.s_ij),
			"\n"
			)




def calculate(problem_param, out_folder):
	problem = LpProblem("Transportation_Problem", LpMinimize)

	# Decision variables
	P_ijt = LpVariable.dicts("P_ijt", indices=((i,j,t) for i in problem_param.product_groups for j in problem_param.factories for t in problem_param.time_periods), cat=const.LpContinuous, lowBound=0)
	R_jt = LpVariable.dicts("R_jt", indices=((j,t) for j in problem_param.factories for t in problem_param.time_periods), cat=const.LpContinuous, lowBound=0)
	IR_jt = LpVariable.dicts("IR_jt", indices=((j,t) for j in problem_param.factories for t in problem_param.time_periods), cat=const.LpContinuous, lowBound=0)
	F_ijkt = LpVariable.dicts("F_ijkt", indices=((i,j,k,t) for i in problem_param.product_groups for j in problem_param.factories for k in problem_param.factories for t in problem_param.time_periods), cat=const.LpContinuous, lowBound=0)
	BO_ijt = LpVariable.dicts("BO_ijt", indices=((i,j,t) for i in problem_param.product_groups for j in problem_param.factories for t in problem_param.time_periods), cat=const.LpContinuous, lowBound=0)
	IN_ijt = LpVariable.dicts("IN_ijt", indices=((i,j,t) for i in problem_param.product_groups for j in problem_param.factories for t in problem_param.time_periods), cat=const.LpContinuous, lowBound=0)
	out_ijt = LpVariable.dicts("out_ijt", indices=((i,j,t) for i in problem_param.product_groups for j in problem_param.factories for t in problem_param.time_periods), cat=const.LpContinuous, lowBound=0)
	X_ijt = LpVariable.dicts("X_ijt", indices=((i,j,t) for i in problem_param.product_groups for j in problem_param.factories for t in problem_param.time_periods), cat=const.LpBinary)
	LS_ij = LpVariable.dicts("LS_ij", indices=((i,j) for i in problem_param.product_groups for j in problem_param.factories), cat=const.LpContinuous, lowBound=0)
	OTD_jt = LpVariable.dicts("OTD_jt", indices=((j,t) for j in problem_param.factories for t in problem_param.time_periods), cat=const.LpContinuous, lowBound=0, upBound=1)
	
	# Objective function
	production_cost = lpSum([problem_param.P_C_ij[i][j] * P_ijt[i,j,t] for i in problem_param.product_groups for j in problem_param.factories for t in problem_param.time_periods ])
	handling_cost = lpSum([problem_param.HD_C_ij[i][j] * F_ijkt[i,j,k,t] for i in problem_param.product_groups for j in problem_param.factories for k in problem_param.factories for t in problem_param.time_periods ])
	transport_cost = lpSum([problem_param.T_C_jk[j][k] * F_ijkt[i,j,k,t] for i in problem_param.product_groups for j in problem_param.factories for k in problem_param.factories for t in problem_param.time_periods ])
	fixproduction_cost = lpSum([problem_param.Fix_ij[i][j]*X_ijt[i,j,t] for i in problem_param.product_groups for j in problem_param.factories for t in problem_param.time_periods ])
	storage_cost = lpSum([problem_param.s_ij[i][j] * IN_ijt[i,j,t] for i in problem_param.product_groups for j in problem_param.factories for t in problem_param.time_periods ])
	backorder_cost = lpSum([problem_param.b_i[i] * BO_ijt[(i,j,t)]*problem_param.v_ij[i][j] for i in problem_param.product_groups for j in problem_param.factories for t in problem_param.time_periods])
	lostsale = lpSum([problem_param.v_ij[i][j]*BO_ijt[i,j,12] for i in problem_param.product_groups for j in problem_param.factories])
	sale = lpSum([problem_param.v_ij[i][j]*out_ijt[i,j,t] for i in problem_param.product_groups for j in problem_param.factories for t in problem_param.time_periods])

	objective = production_cost + fixproduction_cost+ handling_cost + transport_cost + storage_cost + lostsale + backorder_cost 
	watcher = Watcher()
	watcher(P_ijt, R_jt, IR_jt, F_ijkt, BO_ijt, IN_ijt, out_ijt, X_ijt, LS_ij, OTD_jt, production_cost, handling_cost, transport_cost, fixproduction_cost, storage_cost, backorder_cost, lostsale, sale)

	problem += objective
	# Constraint 1
	# for j in problem_param.factories:
	# 	for t in problem_param.time_periods:
	# 		problem += lpSum([problem_param.r_i[i] * P_ijt[i,j,t] for i in problem_param.product_groups]) <= IR_jt[j,t]

	for j in problem_param.factories:
		for t in problem_param.time_periods:
			if t == 1:
				problem += IR_jt[j,t] == problem_param.IR_j0[j] + R_jt[j,t] - (lpSum([problem_param.r_i[i] * P_ijt[i,j,t] for i in problem_param.product_groups]))
			else:
				problem += IR_jt[j,t] == IR_jt[j,t-1] + R_jt[j,t] - (lpSum([problem_param.r_i[i] * P_ijt[i,j,t] for i in problem_param.product_groups]))

	for t in problem_param.time_periods:
		problem += lpSum([R_jt[j,t] for j in problem_param.factories]) <= problem_param.TR_t[t]

	# Constraint 2
	for i in problem_param.product_groups:
		for j in problem_param.factories:
			for t in problem_param.time_periods:
				problem += P_ijt[i,j,t] <= problem_param.Max_P_ij[i][j] * X_ijt[i,j,t]

	# Constraint 3
	for i in problem_param.product_groups:
		for j in problem_param.factories:
			for t in problem_param.time_periods:
				if t == 1:
					problem += IN_ijt[i,j,t] == problem_param.IN_ij0[i][j]  + P_ijt[i,j,t] - out_ijt[i,j,t] + lpSum([ - F_ijkt[i,j,k,t] + F_ijkt[i,k,j,t] for k in (problem_param.factories - set(j))])
				else:
					problem += IN_ijt[i,j,t] == IN_ijt[i,j,t - 1]  + P_ijt[i,j,t] - out_ijt[i,j,t] + lpSum([ - F_ijkt[i,j,k,t] + F_ijkt[i,k,j,t] for k in (problem_param.factories - set(j))])
 
	# Constraint 4
	for i in problem_param.product_groups:
		for j in problem_param.factories:
			for t in problem_param.time_periods:
				if t == 1:
					problem += BO_ijt[i,j,t] == problem_param.D_tij[t][i][j] + problem_param.BO_ij0[i][j] - out_ijt[i,j,t]
				else:
					problem += BO_ijt[i,j,t] == problem_param.D_tij[t][i][j] + BO_ijt[i,j,t-1] - out_ijt[i,j,t]

	# for t in problem_param.time_periods:
	# 			problem += lpSum([P_ijt[i,j,t] for i in problem_param.product_groups for j in problem_param.factories]) >= lpSum([P_ijt[i,j,t] for i in problem_param.product_groups for j in problem_param.factories])

	# #Constraint 5
	# for t in problem_param.time_periods:
	# 	for j in problem_param.factories:
	# 		problem += OTD_jt[j,t] == (lpSum([P_ijt[i,j,t] - BO_ijt[i,j,t] for i in problem_param.product_groups])/((lpSum([problem_param.D_tij[t][i][j] for i in problem_param.product_groups]))+0.01))

	# for t in problem_param.time_periods:
	# 	for j in problem_param.factories:
	# 		problem += OTD_jt[j,t] >= 0.90
	
	# solver = SCIP_PY(msg=False, gapRel=0.0)
	# problem.solve(solver=solver)
	problem.solve()

	status = LpStatus[problem.status]
	watcher(status)
	print("Status:", LpStatus[problem.status])
	print("Optimal Solution:")
	# for v in problem.variables():
	# 	print(v.name, "=", v.varValue)
	print("production_cost :", production_cost.value())
	print("handling_cost :", handling_cost.value())
	print("transport_cost :", transport_cost.value())
	print("storage_cost :", storage_cost.value())
	print("backorder_cost :", backorder_cost.value())
	print("fixproduction_cost :", fixproduction_cost.value())
	print("lostsale_cost :", lostsale.value())
	print("sale :", sale.value())
	print("Optimal Value (Objective Function):", value(problem.objective))
	
	wb2d = openpyxl.Workbook()
	wb2d.create_sheet("R_jt")
	wb2d.create_sheet("IR_jt")
	wb2d.create_sheet("LS_ij")
	wb2d.create_sheet("OTD_jt")

	populate_table(wb2d["R_jt"], (1,1), R_jt, problem_param.factories, problem_param.time_periods)
	populate_table(wb2d["IR_jt"], (1,1), IR_jt, problem_param.factories, problem_param.time_periods)
	populate_table(wb2d["LS_ij"], (1,1), LS_ij, problem_param.product_groups, problem_param.factories)
	populate_table(wb2d["OTD_jt"], (1,1), OTD_jt, problem_param.factories, problem_param.time_periods)

	wb2d.save(out_folder+"/2d.xlsx")

	wb3d = openpyxl.Workbook()
	wb3d.create_sheet("P_ijt")
	populate_table3d(wb3d["P_ijt"], (1,1), P_ijt, problem_param.product_groups, problem_param.factories, problem_param.time_periods)
	wb3d.create_sheet("BO_ijt")
	populate_table3d(wb3d["BO_ijt"], (1,1), BO_ijt, problem_param.product_groups, problem_param.factories, problem_param.time_periods)
	wb3d.create_sheet("IN_ijt")
	populate_table3d(wb3d["IN_ijt"], (1,1), IN_ijt, problem_param.product_groups, problem_param.factories, problem_param.time_periods)
	wb3d.create_sheet("out_ijt")
	populate_table3d(wb3d["out_ijt"], (1,1), out_ijt, problem_param.product_groups, problem_param.factories, problem_param.time_periods)
	wb3d.create_sheet("X_ijt")
	populate_table3d(wb3d["X_ijt"], (1,1), X_ijt, problem_param.product_groups, problem_param.factories, problem_param.time_periods)


	wb3d.create_sheet("TEST_ijt")
	_data = {(i, j, t) : BO_ijt[i,j,t]*problem_param.v_ij[i][j] for i in problem_param.product_groups for j in problem_param.factories for t in problem_param.time_periods}
	populate_table3d(wb3d["TEST_ijt"], (1,1), _data , problem_param.product_groups, problem_param.factories, problem_param.time_periods)
	
	wb3d.save(out_folder+"/3d.xlsx")


	F_ijkt_wb = openpyxl.Workbook()
	for t in problem_param.time_periods:
		sheetname = "F_ijk({})".format(t)
		F_ijkt_wb.create_sheet(sheetname)
		_data = {(j, k, i) : F_ijkt[(i,j,k,t)] for i in problem_param.product_groups for j in problem_param.factories for k in problem_param.factories}
		populate_table3d(F_ijkt_wb[sheetname], (1,1), _data, problem_param.factories, problem_param.factories, problem_param.product_groups)
	F_ijkt_wb.save(out_folder+"/F_ijkt.xlsx")
	return watcher


def populate_table3d(sheet, offset, data, _1_labels, _2_labels, _3_labels):
	col, row = offset
	for _3 in _3_labels:
		sheet.cell(row=row, column=col, value="[{}]".format(_3))
		row += 1
		_data = { (_1, _2) : data[(_1,_2,_3)] for _1 in _1_labels for _2 in _2_labels}
		col, row = populate_table(sheet, (col, row), _data, _1_labels, _2_labels)

		col, _ = offset
		row += 3

def populate_table(sheet, offset, data, col_labels, row_labels):
	col, row = offset

	sheet.cell(row=row, column=0+col, value="")
	_col = col + 1
	for col_lable in col_labels:
		sheet.cell(row=row, column=_col, value=col_lable)
		_col+=1

	_row = row + 1
	for row_label in row_labels:
		_col = col
		sheet.cell(row=_row, column=_col, value=row_label)
		_col += 1
		for col_label in col_labels:
			value = data[(col_label, row_label)].value()
			val = str(value) if value is not None else ""
			sheet.cell(row=_row, column=_col, value=val)
			_col += 1
		_row += 1
	return _col, _row




def read_exel(excel_path):
	workbook = load_workbook(filename=excel_path, data_only=True)
	ret = {}
	for sheet_name in workbook.sheetnames:
		sheet = workbook[sheet_name]
		data = [[cell if cell is not None else "None"  for cell in row] for row in sheet.iter_rows(values_only=True)]
		np_array = np.array(data)
		ret[sheet_name] = np_array
	return ret

def problem_param_generator(problem_param, task):
	copy_of_problem_param = ProblemParam()
	# Boring copy work
	copy_of_problem_param.product_groups = copy.deepcopy(problem_param.product_groups)
	copy_of_problem_param.factories = copy.deepcopy(problem_param.factories)
	copy_of_problem_param.time_periods = copy.deepcopy(problem_param.time_periods)
	copy_of_problem_param.r_i = copy.deepcopy(problem_param.r_i)
	copy_of_problem_param.b_i = copy.deepcopy(problem_param.b_i)
	copy_of_problem_param.TR_t = copy.deepcopy(problem_param.TR_t)
	copy_of_problem_param.IR_j0 = copy.deepcopy(problem_param.IR_j0)
	copy_of_problem_param.Max_P_ij = copy.deepcopy(problem_param.Max_P_ij)
	copy_of_problem_param.IN_ij0 = copy.deepcopy(problem_param.IN_ij0)
	copy_of_problem_param.BO_ij0 = copy.deepcopy(problem_param.BO_ij0)
	# copy_of_problem_param.x_ij = copy.deepcopy(problem_param.x_ij)
	copy_of_problem_param.HD_C_ij = copy.deepcopy(problem_param.HD_C_ij)
	copy_of_problem_param.P_C_ij = copy.deepcopy(problem_param.P_C_ij)
	copy_of_problem_param.T_C_jk = copy.deepcopy(problem_param.T_C_jk)
	copy_of_problem_param.D_tij = copy.deepcopy(problem_param.D_tij)
	copy_of_problem_param.Fix_ij = copy.deepcopy(problem_param.Fix_ij)
	copy_of_problem_param.v_ij = copy.deepcopy(problem_param.v_ij)
	copy_of_problem_param.s_ij = copy.deepcopy(problem_param.s_ij)

	try:
		ref_param = getattr(copy_of_problem_param, task["param"])
	except Exception:
		ref_param = {1:1}

	def gen_add(d, k, l):
		d[k] += task["arg"]

	def gen_more(d, k, l):
		d[k] += d[k] * task["arg"]

	def gen_inc(d, k, l):
		x = getattr(problem_param, task["param"])
		for i in l:
			x = x[i]
		d[k] += task["arg"] * x[k]
		print(x[k],task["arg"],"bamcaigi")

	while True:
		if task["type"] == "+":
			iterate_dict(ref_param, gen_add)
		elif task["type"] == "more":
			iterate_dict(ref_param, gen_more)
		elif task["type"] == "increase":
			iterate_dict(ref_param, gen_inc)
		yield copy_of_problem_param
		

def iterate_dict(d, func, l = []):
	for k in d.keys():
		if callable(getattr(d[k], "keys", None)):
			new_l = list(l)
			new_l.append(k)
			iterate_dict(d[k], func, l = new_l)
		else:
			func(d, k, l)



def generate_iteration_report(watch, filepath):
	watch_set = set()
	for _, data in watch.items():
		watch_set.update(set(data.keys()))

	data = {}
	for w in watch_set:
		if isinstance(watch[0][w], LpAffineExpression):
			data.update({ (iteration, w) : watch[iteration][w] for iteration in watch.keys()})
		elif isinstance(watch[0][w], dict):
			data.update({ (iteration, w) : _Value(sum([var.value() if var.value() is not None else 0  for var in watch[iteration][w].values()])) for iteration in watch.keys()})
		else:
			data.update({ (iteration, w) : _Value(watch[iteration][w]) for iteration in watch.keys() })

	wb = openpyxl.Workbook()
	wb.create_sheet("watcher")
	populate_table(wb["watcher"], (1,1), data, watch.keys(), watch_set)

	# c1 = openpyxl.chart.LineChart()
	# c1.title = "Line Chart"
	# c1.legend = None
	# c1.style = 15
	# c1.y_axis.title = 'Iteration'
	# c1.x_axis.title = 'value'
	# data_excel = openpyxl.chart.Reference(wb["watcher"], min_col=1, min_row=1, max_col=len(watch.keys()), max_row=len(watch_set))
	# c1.add_data(data_excel, titles_from_data=True)
	# wb["watcher"].add_chart(c1, row=1, column=(len(watch_set) + 2))
	
	wb.save(filepath)

def main():
	parser = argparse.ArgumentParser(description="Read Excel file and print its path")
	parser.add_argument('excel_file', type=str, help="Path to the Excel file")
	parser.add_argument('-s', '--sensitive_config', type=str, help="Path to the sensitive config json")
	args = parser.parse_args()
	out_folder = Path(args.excel_file).parent / Path(args.excel_file).stem
	# parse excel
	input_values = read_exel(args.excel_file)

	problem_param = ProblemParam()
	problem_param.parse(input_values)
	problem_param.parameterCorection()
	print(problem_param)

	if args.sensitive_config is None:
		Path(out_folder).mkdir(parents=True, exist_ok=True)
		calculate(problem_param, str(out_folder))
	else:
		with open(args.sensitive_config) as f:
			sensitive_config = json.load(f)
			for task in sensitive_config:
				if not task["active"]:
					continue

				generator = problem_param_generator(problem_param, task)
				watch = {}
				for iteration in range(task["iteration_count"]):
					iteration_out_folder = out_folder / "iteration_{}_{}".format(task["name"], iteration)
					Path(iteration_out_folder).mkdir(parents=True, exist_ok=True)
					Pparame = next(generator)
					watcher = calculate(Pparame, str(iteration_out_folder))
					# print(Pparame)
					Pparame.toExcel(str(iteration_out_folder/ "data.xlsx"))

					if "watch_list" not in task.keys():
						continue

					watch[iteration] = {}
					for w in task["watch_list"]:
						try:
							watch[iteration][w] = watcher[w]
						except Exception as e:
							print("No {} to watch".format(w))
					watch[iteration]["status"] = watcher["status"]


				generate_iteration_report(watch, str(out_folder/"sensitive_report_{}.xlsx".format(task["name"])))


if __name__ == '__main__':
	main()